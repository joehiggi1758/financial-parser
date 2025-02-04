import os
import warnings
import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

# ----------------------------------------------------------------
# Configure warnings and logging
warnings.simplefilter("ignore", category=UserWarning)
warnings.simplefilter("ignore", category=DeprecationWarning)

logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)

# ----------------------------------------------------------------
def extract_metadata(file_path: str, sheet_name: str) -> dict:
    """
    Capture workbook & sheet info, PLUS any meta fields from 
    specific cells at the top of the sheet.

    Adjust the cell references below to match your layout.
    For example, we read:
       - Company from cell B2
       - Segment from cell B3
       - Period End from cell B4

    If your info is in different cells, just change them.
    """
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {
            "workbook_name": Path(file_path).stem,
            "sheet_name": sheet_name,
            "company": None,
            "segment": None,
            "period_end": None
        }

    sheet = wb[sheet_name]

    # Example reads from B2, B3, B4
    # If your file has different references, adjust these
    company_val = sheet["B2"].value
    segment_val = sheet["B3"].value
    period_val = sheet["B4"].value

    return {
        "workbook_name": Path(file_path).stem,
        "sheet_name": sheet_name,
        "company": company_val,
        "segment": segment_val,
        "period_end": period_val
    }

# ----------------------------------------------------------------
def build_row_data(df: pd.DataFrame) -> list:
    """
    Convert each row of DataFrame 'df' into a dictionary with:
      - row index or 'excel_row'
      - category (e.g. 'IS Statement Header', 'Sub Header', etc.)
      - all column data from the row

    This example uses the first column (df.columns[0]) to decide category:
      - If the text has 'header' (case-insensitive) => 'IS Statement Header'
      - If the text has 'sub' => 'Sub Header'
      - Else => 'Financial Metric'

    Adjust if you need a different column or logic.
    """
    row_data = []
    if df.empty:
        return row_data

    # We’ll use the first column to detect headers
    header_col = df.columns[0]

    for i, row_series in df.iterrows():
        raw_text = str(row_series[header_col])  # text in the first column
        text_lower = raw_text.strip().lower()

        if "header" in text_lower:
            category = "IS Statement Header"
        elif "sub" in text_lower:
            category = "Sub Header"
        else:
            category = "Financial Metric"

        row_dict = {
            "excel_row": i + 1,           # Excel row index
            "category": category,
            "values": row_series.to_dict()  # store all cell data from that row
        }
        row_data.append(row_dict)

    return row_data

# ----------------------------------------------------------------
def assign_headers(row_data: list) -> list:
    """
    Traverse 'row_data' in reverse, so we can propagate the actual text 
    from each 'IS Statement Header' or 'Sub Header' row to the rows below it.

    We'll store the exact text from that row in row["IS Statement Header"] 
    or row["Sub Header"]. If it's just a normal row, we can store the text
    as "Main Header" (optional).

    By uncommenting the lines at the end, we ensure that every row 
    inherits the nearest header or sub-header from above.
    """
    current_is_header_text = None
    current_sub_header_text = None
    current_main_header_text = None

    if not row_data:
        return row_data

    for row in reversed(row_data):
        cat = row["category"]
        # The entire row's data is in row["values"]
        # We'll use the same first column we used for detection
        all_values = row["values"]
        first_col_name = list(all_values.keys())[0]  # or df.columns[0]
        row_text = str(all_values[first_col_name]).strip()

        if cat == "IS Statement Header":
            current_is_header_text = row_text
            row["IS Statement Header"] = current_is_header_text
        elif cat == "Sub Header":
            current_sub_header_text = row_text
            row["Sub Header"] = current_sub_header_text
        else:
            current_main_header_text = row_text
            row["Main Header"] = current_main_header_text

        # Propagate these to *every* row below:
        row["IS Statement Header"] = current_is_header_text
        row["Sub Header"] = current_sub_header_text
        row["Main Header"] = current_main_header_text

    return row_data

# ----------------------------------------------------------------
def build_cleaned_dataframe(row_data: list) -> pd.DataFrame:
    """
    Convert the list of dictionaries (row_data) into a DataFrame,
    preserving all columns from the row plus:
      - "Excel Row"
      - "Category"
      - "IS Statement Header"
      - "Sub Header"
      - "Main Header"
    """
    if not row_data:
        return pd.DataFrame()

    flattened = []
    for rd in row_data:
        base = {
            "Excel Row": rd["excel_row"],
            "Category": rd["category"],
            "IS Statement Header": rd.get("IS Statement Header"),
            "Sub Header": rd.get("Sub Header"),
            "Main Header": rd.get("Main Header"),
        }
        # Merge in original row columns
        for k, v in rd["values"].items():
            base[str(k)] = v
        flattened.append(base)

    return pd.DataFrame(flattened)

# ----------------------------------------------------------------
def process_sheet(file_path: str, sheet_name: str, skip_header_rows=1) -> pd.DataFrame:
    """
    Load a single sheet, extract rows, assign headers, return final DataFrame.

    By default, skip_header_rows=1 => The *first* row in the sheet is treated 
    as column headers. If your sheet has no header row, set skip_header_rows=0 
    so the entire sheet is data.

    Steps:
      1. Convert the sheet to a DataFrame.
      2. Detect categories & build row_data.
      3. Propagate "IS Statement Header" or "Sub Header" text downward.
      4. Build cleaned DataFrame with all columns + new fields.
      5. Print/log any "IS Statement Header" rows found.
    """
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        logging.warning(f"Sheet '{sheet_name}' not found in '{file_path}'")
        return pd.DataFrame()

    sheet = wb[sheet_name]
    data_rows = list(sheet.values)

    if not data_rows:
        return pd.DataFrame()

    # If the first row has column names, parse them:
    if skip_header_rows > 0:
        columns = data_rows[0]  # The first row is columns
        data_rows = data_rows[skip_header_rows:]  # skip them from data
        df = pd.DataFrame(data_rows, columns=columns)
    else:
        # No header rows, so all rows are data with numeric columns
        df = pd.DataFrame(data_rows)

    # Build structured rows
    row_data = build_row_data(df)

    # Assign headers (IS Statement Header, Sub Header, etc.) in reverse
    row_data = assign_headers(row_data)

    # Convert row_data into a cleaned DataFrame
    cleaned_df = build_cleaned_dataframe(row_data)

    # Print any "IS Statement Header" rows for debugging
    is_mask = cleaned_df["Category"] == "IS Statement Header"
    if is_mask.any():
        logging.info(f"Workbook: {file_path}, Sheet: {sheet_name} — Found IS Statement Header(s):")
        for idx, row in cleaned_df[is_mask].iterrows():
            logging.info(f"  Excel Row {row['Excel Row']}: '{row['IS Statement Header']}'")

    return cleaned_df

# ----------------------------------------------------------------
def process_workbooks(input_path: str, output_file: str, skip_header_rows=1) -> None:
    """
    Iterate over .xls/.xlsx files in 'input_path', process each sheet,
    gather the data into one DataFrame, write to 'output_file'.

    We also attach:
      - Workbook Name
      - Sheet Name
      - Company
      - Segment
      - Period End
    to every row, by calling 'extract_metadata(...)' for each sheet.
    """
    all_data = []
    files = list(Path(input_path).glob("*.xls*"))  # matches .xls or .xlsx
    if not files:
        logging.warning("No Excel files found in input path.")

    for file_path in tqdm(files, desc="Processing Excel files", unit="file"):
        # For each file, load it and read the metadata for each sheet
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            # Extract meta info (like Company, Segment, Period End)
            meta = extract_metadata(str(file_path), sheet_name)

            # Now process the actual data from the sheet
            df = process_sheet(str(file_path), sheet_name, skip_header_rows=skip_header_rows)
            if not df.empty:
                # Attach the meta columns
                df["Workbook Name"] = meta["workbook_name"]
                df["Sheet Name"]    = meta["sheet_name"]
                df["Company"]       = meta["company"]
                df["Segment"]       = meta["segment"]
                df["Period End"]    = meta["period_end"]

                all_data.append(df)

    # Once we've processed all files & sheets, combine them
    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        combined_df.drop_duplicates(inplace=True)
        combined_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        logging.info(f"Combined output saved to: {output_file}")
    else:
        logging.info("No data found in any workbook; no output generated.")

# ----------------------------------------------------------------
if __name__ == "__main__":
    # Example usage:
    input_path = "data/input"            # Folder with your .xls/.xlsx files
    output_file = "data/output/test.csv"

    # If your sheet has exactly 1 header row (the first row are column names),
    # use skip_header_rows=1. 
    # If there's no header row, set skip_header_rows=0.
    process_workbooks(input_path, output_file, skip_header_rows=1)
    logging.info("All processing complete!")
